#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""extract.py — Extrait un snapshot JSON depuis l'Excel de suivi.

Usage :
    python3 extract.py                          # snapshot daté d'aujourd'hui
    python3 extract.py --date 2026-07-06 --label "Migration initiale"

Le snapshot est écrit dans data/snapshots/<date>.json (niveau 3 — jamais publié).
"""
import argparse, json, os, sys, datetime
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = "/Users/david/Downloads/OAQ-Suivi-plan-strategique-SIMPLE.xlsx"

def extract(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sv = wb["Suivi"]
    actions = []
    for r in range(2, sv.max_row + 1):
        no = sv.cell(r, 1).value
        if not no:
            continue
        no = str(no).strip()
        star = no.startswith("★")
        actions.append({
            "no": no.replace("★", "").strip(),
            "star": star,
            "act": str(sv.cell(r, 2).value or "").strip(),
            "qui": str(sv.cell(r, 3).value or "").strip(),
            "statut": str(sv.cell(r, 4).value or "").strip(),
            "bloque": str(sv.cell(r, 5).value or "Non").strip() == "Oui",
            "quoi": str(sv.cell(r, 6).value or "").strip(),
            "maj": str(sv.cell(r, 7).value or "").strip(),
        })
    ja = wb["Jalons"]
    jalons = []
    for r in range(2, ja.max_row + 1):
        p = ja.cell(r, 1).value
        if not p:
            continue
        p = str(p).strip()
        jalons.append({
            "prio": p.replace("★", "").strip(),
            "star": p.startswith("★"),
            "jalon": str(ja.cell(r, 2).value or "").strip(),
            "qui": str(ja.cell(r, 3).value or "").strip(),
            "echeance": str(ja.cell(r, 4).value or "").strip(),
            "statut": str(ja.cell(r, 5).value or "").strip(),
            "comm": str(ja.cell(r, 6).value or "").strip(),
        })
    kp = wb["KPIs"]
    kpis = []
    for r in range(2, kp.max_row + 1):
        k = kp.cell(r, 1).value
        if not k:
            continue
        kpis.append({
            "id": str(k).strip(),
            "kpi": str(kp.cell(r, 2).value or "").strip(),
            "baseline": str(kp.cell(r, 3).value or "").strip(),
            "actuel": str(kp.cell(r, 4).value or "").strip(),
            "cible": str(kp.cell(r, 5).value or "").strip(),
            "comm": str(kp.cell(r, 6).value or "").strip(),
        })
    return actions, jalons, kpis

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", default=datetime.date.today().isoformat())
    ap.add_argument("--label", default="")
    ap.add_argument("--xlsx", default=XLSX)
    a = ap.parse_args()
    actions, jalons, kpis = extract(a.xlsx)
    snap = {"date": a.date, "label": a.label or f"État au {a.date}", "demo": False,
            "actions": actions, "jalons": jalons, "kpis": kpis}
    out = os.path.join(BASE, "data", "snapshots", f"{a.date}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=1)
    print(f"Snapshot écrit : {out} ({len(actions)} actions, {len(jalons)} jalons, {len(kpis)} KPIs)")

if __name__ == "__main__":
    main()
