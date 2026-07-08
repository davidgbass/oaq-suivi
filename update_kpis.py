#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""update_kpis.py v2 — Reconstruit l'onglet KPIs (directives présidence 2026-07-08) :
  · Colonne « Tendance (3 exercices) » ajoutée
  · K3 = nouveaux permis avec tendance 36→33→23→16
  · K1 détaillé (sous-mesures), K2 mieux expliqué, K4 à partager avec la CIP
  · K5/K6 binaires tel quel · K7 EN SUSPENS (exercice de recension en cours)
  · C1-C3 indicateurs de contexte

Usage : python3 update_kpis.py <source.xlsx> <sortie.xlsx>
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

SRC, OUT = sys.argv[1], sys.argv[2]
wb = openpyxl.load_workbook(SRC)
INK = "1F2933"; TEAL = "00627A"; GREY = "7d8f96"
thin = Side(style="thin", color="D9D9D9"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
topleft = Alignment(horizontal="left", vertical="top", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# repartir d'un onglet propre
if "KPIs" in wb.sheetnames:
    del wb["KPIs"]
kp = wb.create_sheet("KPIs", 2)
kp.sheet_view.showGridLines = False
kp.sheet_properties.tabColor = "2E7D8A"

HEAD = ["#", "KPI", "Baseline 2025-2026", "Valeur actuelle",
        "Tendance (3 exercices)", "Cibles (à adopter sept. 2026)", "Commentaire"]
kp.append(HEAD)
for c in range(1, len(HEAD) + 1):
    cell = kp.cell(1, c)
    cell.font = Font(b=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1B5F70")
    cell.alignment = center; cell.border = border
kp.row_dimensions[1].height = 30

ROWS = [
 ("K1", "Satisfaction globale des membres (/10)",
  "8,5/10 (n=381 — congrès 2025)", "",
  "— (1re mesure ; prochaine : congrès 2027)",
  "",
  "Q1 Slido : « Quel est votre niveau de satisfaction globale envers le travail de l'OAQ ? » (84 % cotent 8+). "
  "DÉTAIL des sous-mesures 2025 : partenaire qui m'encadre et me soutient 8,6 · outils de conformité 8,6 · "
  "équité et inclusion 9,3 · utilité de la formation continue 96 % d'accord. À détailler davantage au sondage 2027 (instrument à raffiner)."),
 ("K2", "Perception de la représentation politique (%)",
  "39 % (n=160 — Qualtrics Q13, congrès 2025)", "",
  "— (1re mesure ; prochaine : congrès 2027)",
  "",
  "CE QUE ÇA MESURE : le % de membres qui reconnaissent que l'Ordre « représente les intérêts des membres auprès des instances "
  "gouvernementales ». POURQUOI C'EST LE KPI DE L'ORIENTATION 3 : c'est l'action la MOINS reconnue de toutes (39 % vs surveillance 84 %, "
  "formations 78 %, normes 77 % — moyenne 63 %). Si le plan réussit (lobbying RAMQ, présence politique), ce chiffre doit monter."),
 ("K3", "Nouveaux permis délivrés par année (relève)",
  "16 (2024-25 : 14 art. 184 + 2 hors Canada)", "",
  "36 → 33 → 23 → 16  ↘  (2021-22 → 2024-25)",
  "",
  "SUBSTITUTION PROPOSÉE (à entériner CA sept. 2026) : l'effectif total n'a jamais baissé (277→548 depuis 2010) ; "
  "c'est le flux d'entrée qui s'effondre (−64 % depuis 2018-19 : 45→16). Série exacte 15 ans (rapports annuels). "
  "Sensible aux actions relève 1.1.x."),
 ("K4", "Conformité à l'inspection — % des dossiers sans recommandation",
  "À définir avec la CIP (nouvelle série 2025-26)", "",
  "— (série à créer)",
  "",
  "À PARTAGER avec le comité d'inspection professionnelle et sa responsable : aucun « taux de conformité » n'a jamais été publié ; "
  "l'historique est inextractible (3 régimes de mesure : 2010-15, 2015-20, pilote 2021-). Demande à la CIP : produire le dénominateur "
  "« membres inspectés avec ≥ 1 recommandation ». Repli immédiat : taux de couverture (2024-25 : 112/548 ≈ 20,4 % — cible 20 %/an)."),
 ("K5", "Bonification RAMQ obtenue (oui/non)",
  "NON", "",
  "—  (binaire : obtenu ou pas)",
  "—",
  "Tel quel, ON/OFF (décision présidence 2026-07-08). Jalons documentés : étude AppEco (8,90 $/$), lettre ouverte + coalition "
  "(10 déc. 2025), rencontre ministre (déc. 2025), mémoire PL-15 (12 févr. 2026). Jalon P1.2 : mémoire RAMQ d'ici déc. 2026."),
 ("K6", "Accord d'alignement OOAQ signé (oui/non)",
  "NON", "",
  "—  (binaire : obtenu ou pas)",
  "—",
  "Tel quel, ON/OFF (décision présidence 2026-07-08). Véhicule : comité interordres actes professionnels en santé auditive "
  "(OAQ+OOAQ+CMQ, mandaté CA 2025-06-06). Représentant OAQ à nommer. Canal actif (rencontre 18 déc. 2025)."),
 ("K7", "Présence stratégique (nb d'instances avec mandat actif)",
  "EN SUSPENS", "",
  "Exercice de vérification en cours",
  "",
  "EN SUSPENS (décision présidence 2026-07-08) : difficile à calculer tel que formulé. Exercice lancé : recension de TOUS les rapports "
  "d'activité de la présidence 2023-2026 (PV des CA) pour vérifier si l'augmentation de la représentation est réelle et documentable. "
  "Liste candidate provisoire : CIQ · comité interordres · comité OTC-APAQ · CFOR · coalition RAMQ."),
 ("C1", "Contexte — Effectif total au tableau",
  "548 (31 mars 2025)", "", "↗ constant (277→548 depuis 2010, +3 en 2024-25)", "—",
  "Sans cible : le stock croît encore, mais le ralentissement est net — contexte du KPI relève K3."),
 ("C2", "Contexte — Résultat financier annuel · actifs nets",
  "−406 202 $ · 1 179 659 $ (2024-25)", "", "actifs nets : 1 586 k$ → 1 180 k$  ↘", "—",
  "Sans cible : surveillance CAUD. Affiché parce que la pérennité opérationnelle est l'orientation 4 du plan."),
 ("C3", "Contexte — Participants au congrès annuel",
  "≈ 500 (2024-25)", "", "≈ 500 stable (499 en 2023-24)", "—",
  "Sans cible : série de 16 ans (~400 → ~500). Signal d'engagement des membres."),
]

for i, row in enumerate(ROWS, start=2):
    ctx = row[0].startswith("C")
    for c, v in enumerate(row, start=1):
        cell = kp.cell(i, c, v)
        cell.border = border
        cell.font = Font(size=10.5, color=INK, italic=ctx)
        cell.alignment = center if c in (1, 4) else topleft
    kp.cell(i, 1).font = Font(b=True, size=11, color=GREY if ctx else TEAL, italic=ctx)

last = len(ROWS) + 1
for col, w in {"A": 5, "B": 40, "C": 26, "D": 14, "E": 30, "F": 22, "G": 62}.items():
    kp.column_dimensions[col].width = w
kp.freeze_panes = "B2"
t = Table(displayName="tblKPIs", ref=f"A1:G{last}")
t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
kp.add_table(t)

wb.save(OUT)
print(f"OK v2 — onglet KPIs reconstruit (7 KPIs + 3 contexte, colonne Tendance) → {OUT}")
